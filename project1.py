import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('project1.xlsx')
sheet = wb['Sheet1']
cell = sheet['a1']
cell = sheet.cell(row=1, column=1)
print(cell.value)

for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row=row, column=3)
    corrected_price = cell.value * 0.9
    corrected_cell = sheet.cell(row=row, column=4)
    corrected_cell.value = corrected_price

values = Reference(sheet, min_row=2,
          max_row=sheet.max_row,
          min_col=4,        #jitne col main value hai vo yha likhna hai min and max value khase start hua hai aur kha end ho raha hai
          max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('project4.xlsx')